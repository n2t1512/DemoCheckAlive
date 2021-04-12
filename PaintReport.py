import pandas as pd
import openpyxl
import xlsxwriter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.cell import Cell
from openpyxl.styles import colors
from datetime import datetime
import os
from openpyxl.comments import Comment
from datetime import date


def MergeSheets():

    data_frames = []
    today = date.today()
    sheet_name = today.strftime("%d-%b")
    file_name = today.strftime("%b-%y")

    if not os.path.exists("Check-Alive-Lotte-" + str(file_name)+'.xlsx'):
        wb = openpyxl.Workbook()
        wb.save("Check-Alive-Lotte-" + str(file_name)+'.xlsx')

    xl2 = openpyxl.load_workbook("Check-Alive-Lotte-" + str(file_name) + '.xlsx')

    if os.path.exists("B2_Lotte_Apr_FPT.xlsx"):
        df1 = pd.read_excel("B2_Lotte_Apr_FPT.xlsx", sheet_name=str(sheet_name), skiprows=0)
        data_frames.append(df1)

    # if os.path.exists("B2_Lotte_Apr_VT.xlsx"):
    #     df2 = pd.read_excel("B2_Lotte_Apr_VT.xlsx", sheet_name=str(sheet_name), skiprows=0)
    #     values2 = df2[['Viettel']]
    #     data_frames.append(values2)
    #
    # if os.path.exists("B2_Lotte_Apr_VNPT.xlsx"):
    #     df3 = pd.read_excel("B2_Lotte_Apr_VNPT.xlsx", sheet_name=str(sheet_name), skiprows=0)
    #     values3 = df3[['VNPT']]
    #     data_frames.append(values3)
    #
    # if os.path.exists("B2_Lotte_Apr_Mobi.xlsx"):
    #     df4 = pd.read_excel("B2_Lotte_Apr_Mobi.xlsx", sheet_name=str(sheet_name), skiprows=0)
    #     values4 = df4[['Mobi 4G']]
    #     data_frames.append(values4)

    join = pd.concat(data_frames, axis=1, ignore_index=False)
    with pd.ExcelWriter("Check-Alive-Lotte-" + str(file_name)+".xlsx", sheet_name=str(sheet_name), engine='openpyxl', mode='a') as writer:
        join.to_excel(writer, sheet_name=str(sheet_name))
    writer.save()

    # writer = pd.ExcelWriter("Check-Alive-Lotte-" + str(file_name) + '.xlsx')
    # join.to_excel("Check-Alive-Lotte-" + str(file_name)+".xlsx", sheet_name=str(sheet_name), startrow=1)


def ToMau():

    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    green_fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')
    orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    random_fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    mid = Alignment(horizontal='center')
    font_bold = Font(bold=True)

    today = date.today()
    sheet_name = today.strftime("%d-%b")
    file_name = today.strftime("%b-%y")

    #xl2 = openpyxl.load_workbook("Check-Alive-Lotte-" + str(file_name)+'.xlsx')
    xl2 = openpyxl.load_workbook("B2_Lotte_Apr_FPT.xlsx")
    sheet2 = xl2[str(sheet_name)]

    # sheet2['A1'] = "Check Alive"
    # sheet2['A1'].fill = random_fill
    # sheet2['A1'].border = thin_border
    # sheet2['A1'].alignment = mid
    # sheet2['A1'].font = Font(bold=True, size=20)
    # sheet2.merge_cells(start_row=1, end_row=1, start_column=1, end_column=8)
    sheet2.column_dimensions['A'].width = 10
    sheet2.column_dimensions['B'].width = 20
    sheet2.column_dimensions['C'].width = 40
    sheet2.column_dimensions['D'].width = 20
    sheet2.column_dimensions['E'].width = 20
    sheet2.column_dimensions['F'].width = 20
    sheet2.column_dimensions['G'].width = 20
    sheet2.column_dimensions['H'].width = 20

    so_dong = sheet2.max_row
    so_cot = sheet2.max_column

    for i in range(1, so_dong + 1):
        for j in range(1, so_cot + 1):
            sheet2.cell(i, j).border = thin_border

    range_selected = []
    for i in range(1, so_dong + 1):
        for j in range(4, so_cot + 1):
            range_selected.append(sheet2.cell(row=i, column=j))

    for i in range_selected:
        if i.value == "PASSED":
            i.fill = green_fill
        elif i.value == None:
            i.fill = red_fill
        elif i.value == "Viettel" or i.value == "VNPT" or i.value == "FPT" or i.value == "Mobi 4G":
            i.fill = orange_fill
            i.font = font_bold
            i.alignment = mid
        else:
            i.fill = red_fill
            x = i.value.rsplit("FAILED")
            comment = Comment(x, "Comment Author")
            i.comment = comment
            i.value = "FAILED"

    range_selected = []
    for i in range(1, so_dong + 1):
        for j in range(1, so_cot + 1):
            range_selected.append(sheet2.cell(row=i, column=j))

    for i in range_selected:
        if i.value == "Time" or i.value == "Domain" or i.value == "Functions":
            i.fill = orange_fill
            i.font = font_bold
            i.alignment = mid
    xl2.save("B2_Lotte_Apr_FPT.xlsx")
    #xl2.save("Check-Alive-Lotte-" + str(file_name)+'.xlsx')


if __name__ == '__main__':
    # MergeSheets()
    ToMau()

